import openpyxl
from openpyxl import Workbook,load_workbook
from datetime import date

today = date.today()

date_string = today.strftime("%d-%m-%Y")


class PufExcell:
    def __init__(self,file_path='puffstock.xlsx'):
        self.file_path = file_path
        self.book = load_workbook(filename=file_path)
        #self.sheet = self.book.create_sheet(title=date_string) if date_string not in self.book.sheetnames else self.book[date_string]#create new sheet every day
        self.sheet = self.book.active
        self.orders = [] 
        self.row_value =None
    def save_workbook(self):
        self.book.save(self.file_path)

    
    def add_order(self,*args):
        
        if self.sheet.max_row == 1:
            self.sheet.append(('SİP_NO','İSİM', 'TELEFON NO', 'ÜRÜN', 'ADET','ÖDEME DURUMU','NOT','SİPARİŞ TARİHİ'))        
        
        
        for i in range(self.sheet.max_row, self.sheet.max_row+1):
            
            self.sheet.append(([i] + list(args)+[date_string]))

        self.save_workbook()
     

    
    def find_order_excel(self,order):
        
         
        if len(self.orders)==0:
            for row in self.sheet.iter_rows(min_row=1, min_col=1, max_row=None, max_col=2):
                
                temp_orders=[]
                for cell in row:  
                    
                    if cell.value == order:
                        
                        
                        for c in self.sheet.iter_cols(min_row=cell.row, max_row=cell.row):
                            
                            for r in c:
                                
                                
                                temp_orders.append(r.value)
                            
                        self.orders.append(temp_orders)
                            
        self.save_workbook()
       
        
        return self.orders
    def saving_private_row(self, row_value):

        self.row_value=row_value
        return self.row_value

    def update_order_excel(self, *args):
     
        # print(arg[idx])
      
        for idx,arg in enumerate(args,start=2):
            
            self.sheet.cell(row=self.row_value, column=idx).value = arg
        self.save_workbook()
        
     
    def dalete_order_excel(self, *args):
        for idx,arg in enumerate(args,start=2):
            
            self.sheet.cell(row=self.row_value, column=idx).value = arg
        self.save_workbook()
        

        self.save_workbook()
           
