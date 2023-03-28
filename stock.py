import openpyxl
from openpyxl import Workbook,load_workbook
from datetime import date

today = date.today()

date_string = today.strftime("%d-%m-%Y")


class StockExcell:
    def __init__(self,file_path='puffstock.xlsx'):
        self.file_path = file_path
        self.book = load_workbook(filename=file_path)
        self.product = [] 
        if "Stock" in self.book.sheetnames:
            self.s_sheet = self.book["Stock"]
            if self.s_sheet.max_row == 1:
                self.s_sheet.append(('PRO_N0','TÜR', 'ADET'))    

    def save_workbook(self):
        self.book.save(self.file_path)

    
    def add_stock(self, *args):
        args = list(args)
        item_name = args[0]

        # Check if item already exists in the sheet
        for row in self.s_sheet.iter_rows():
            if row[1].value == item_name:
                print(f"{item_name} already exists in the sheet.")
                return

        new_row = [self.s_sheet.max_row + 1] + args
        self.s_sheet.append(new_row)
        print(f"{item_name} has been added to the sheet.")

        self.save_workbook()
    def find_product_excel(self,product):
        
        self.product=[]
        if len(self.product)==0:
            for row in self.s_sheet.iter_rows(min_row=1, min_col=1, max_row=None, max_col=2):
                
                temp_orders=[]
                for cell in row:  
                    
                    if cell.value == product:
                        
                        
                        for c in self.s_sheet.iter_cols(min_row=cell.row, max_row=cell.row):
                            
                            for r in c:
                                
                                
                                temp_orders.append(r.value)
                            
                        self.product.append(temp_orders)
                            
        self.save_workbook()
        return self.product
        # for i in range(self.sheet.max_row, self.sheet.max_row+1):
    
    def update_product_excel(self, *args):
     
        args=list(args)
        
       
        for idx,arg in enumerate(args[1:],start=2):
  
            self.s_sheet.cell(row=self.row_value, column=idx).value = arg     
        
        self.save_workbook()
    def delete_order_excel(self, *args):
        
            
         
        
        self.s_sheet.delete_rows(self.row_value)
        
        
        self.save_workbook()
            